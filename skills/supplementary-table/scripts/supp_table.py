#!/usr/bin/env python3
"""
supplementary-table: write supplementary tables that name the figure they back.

Library API:
    from supp_table import write_csv, write_tsv, write_xlsx

CLI:
    python supp_table.py csv  <out.csv>  --rows-from data.csv [provenance flags…]
    python supp_table.py tsv  <out.tsv>  --rows-from data.tsv [provenance flags…]
    python supp_table.py xlsx <out.xlsx> --sheet NAME=path.csv [...]  [provenance flags…]

Provenance flags (CLI + library kwargs):
    --figure FIGURE_ID       e.g. figure_3_panel_b
    --script SCRIPT_REF      e.g. scripts/build_figures.py::build_fig3
    --source SOURCE_PATH     e.g. runs/2026-05-25/summary.json
    --description TEXT       free-text one-liner

Output:
    csv / tsv  →  `#`-prefixed comment header block, then data rows.
    xlsx       →  data sheet(s) + a `Provenance` sheet.

sha256 is computed over the canonical data block (excluding the comment header
in csv / tsv; excluding the Provenance sheet in xlsx) so the same data hashes
identically across formats.

Stdlib only except for `openpyxl` (xlsx).
"""

from __future__ import annotations

import argparse
import csv
import datetime as _dt
import hashlib
import io
import os
import sys
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Union

__all__ = ["write_csv", "write_tsv", "write_xlsx"]

# ---------- core helpers -----------------------------------------------------


def _stamped_at() -> str:
    return _dt.datetime.now(tz=_dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _canonical_columns(rows: Sequence[Mapping[str, Any]]) -> List[str]:
    """First non-empty row's keys, in insertion order."""
    for row in rows:
        return list(row.keys())
    return []


def _row_block_bytes(rows: Sequence[Mapping[str, Any]], delimiter: str) -> bytes:
    """Serialize just the data rows (no comments, no provenance) for hashing."""
    cols = _canonical_columns(rows)
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=delimiter, lineterminator="\n")
    w.writerow(cols)
    for row in rows:
        w.writerow([_stringify(row.get(c)) for c in cols])
    return buf.getvalue().encode("utf-8")


def _stringify(v: Any) -> str:
    if v is None:
        return ""
    return str(v)


def _provenance_lines(
    *,
    figure: Optional[str],
    script: Optional[str],
    source: Optional[str],
    sha256: str,
    stamped_at: str,
    description: Optional[str],
    comment_prefix: str = "# ",
) -> List[str]:
    out: List[str] = []
    for key, value in (
        ("figure", figure),
        ("script", script),
        ("source", source),
        ("sha256", sha256),
        ("stamped_at", stamped_at),
        ("description", description),
    ):
        if value:
            out.append(f"{comment_prefix}{key}: {value}")
    return out


# ---------- CSV / TSV --------------------------------------------------------


def _write_delimited(
    path: Union[str, os.PathLike],
    rows: Sequence[Mapping[str, Any]],
    *,
    delimiter: str,
    figure: Optional[str] = None,
    script: Optional[str] = None,
    source: Optional[str] = None,
    description: Optional[str] = None,
) -> Dict[str, Any]:
    rows = list(rows)
    data_bytes = _row_block_bytes(rows, delimiter=delimiter)
    sha = hashlib.sha256(data_bytes).hexdigest()
    stamped = _stamped_at()
    prov = _provenance_lines(
        figure=figure,
        script=script,
        source=source,
        sha256=sha,
        stamped_at=stamped,
        description=description,
    )
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        for line in prov:
            fh.write(line + "\n")
        fh.write(data_bytes.decode("utf-8"))
    return {
        "path": str(path),
        "figure": figure,
        "script": script,
        "source": source,
        "sha256": sha,
        "stamped_at": stamped,
        "n_rows": len(rows),
    }


def write_csv(
    path: Union[str, os.PathLike],
    rows: Sequence[Mapping[str, Any]],
    *,
    figure: Optional[str] = None,
    script: Optional[str] = None,
    source: Optional[str] = None,
    description: Optional[str] = None,
) -> Dict[str, Any]:
    """Write a single-sheet CSV with a `#`-prefixed provenance header block."""
    return _write_delimited(
        path,
        rows,
        delimiter=",",
        figure=figure,
        script=script,
        source=source,
        description=description,
    )


def write_tsv(
    path: Union[str, os.PathLike],
    rows: Sequence[Mapping[str, Any]],
    *,
    figure: Optional[str] = None,
    script: Optional[str] = None,
    source: Optional[str] = None,
    description: Optional[str] = None,
) -> Dict[str, Any]:
    """Write a single-sheet TSV with a `#`-prefixed provenance header block."""
    return _write_delimited(
        path,
        rows,
        delimiter="\t",
        figure=figure,
        script=script,
        source=source,
        description=description,
    )


# ---------- XLSX -------------------------------------------------------------


def write_xlsx(
    path: Union[str, os.PathLike],
    sheets: Mapping[str, Sequence[Mapping[str, Any]]],
    *,
    figure: Optional[str] = None,
    script: Optional[str] = None,
    source: Optional[str] = None,
    description: Optional[str] = None,
    per_sheet_figures: Optional[Mapping[str, str]] = None,
    per_sheet_sources: Optional[Mapping[str, str]] = None,
    per_sheet_descriptions: Optional[Mapping[str, str]] = None,
    banner: bool = True,
) -> Dict[str, Any]:
    """
    Write a multi-sheet XLSX. Adds a `Provenance` sheet listing
    sheet_name | figure_id | script | source | sha256 | stamped_at | description
    for every data sheet.

    `figure` / `source` apply to every sheet unless overridden per-sheet via
    `per_sheet_figures` / `per_sheet_sources`. `script` is sheet-agnostic
    (one row of provenance per sheet, with the script column repeating the
    global value). `description` can be global, per-sheet via
    `per_sheet_descriptions`, or both (per-sheet wins).

    When `banner=True` (default), each data sheet gets a bold top row of
    the form `Backs Figure 3` / `Backs Supplementary Figure S2` / etc.
    derived from the sheet's `figure_id`. Set `banner=False` to suppress
    if you want the unadorned header row at row 1.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as e:
        raise RuntimeError(
            "supplementary-table.write_xlsx requires openpyxl. "
            "Install with: pip install openpyxl"
        ) from e

    per_sheet_figures = dict(per_sheet_figures or {})
    per_sheet_sources = dict(per_sheet_sources or {})
    per_sheet_descriptions = dict(per_sheet_descriptions or {})
    stamped = _stamped_at()

    wb = Workbook()
    # remove the default empty sheet
    wb.remove(wb.active)

    # Styling — kept minimal so the skill is opinion-free on the rest of
    # the sheet. Just the banner needs to stand out.
    banner_font = Font(bold=True, color="0f172a", size=11)
    banner_fill = PatternFill("solid", fgColor="fef3c7")
    banner_align = Alignment(horizontal="left", vertical="center")

    per_sheet_records: List[Dict[str, Any]] = []
    for sheet_name, rows in sheets.items():
        rows = list(rows)
        cols = _canonical_columns(rows)
        fig_id = per_sheet_figures.get(sheet_name, figure or "")
        desc   = per_sheet_descriptions.get(sheet_name, description or "")

        ws = wb.create_sheet(title=sheet_name)
        if banner and fig_id:
            banner_text = _format_banner(fig_id)
            ws.append([banner_text])
            cell = ws["A1"]
            cell.font = banner_font
            cell.fill = banner_fill
            cell.alignment = banner_align
            # Span the banner across all data columns so the colour fill
            # reads as a band, not a single cell.
            if cols:
                last_col_letter = _col_letter(len(cols))
                for j in range(2, len(cols) + 1):
                    ws[f"{_col_letter(j)}1"].fill = banner_fill
                ws.merge_cells(f"A1:{last_col_letter}1")
        ws.append(cols)
        for row in rows:
            ws.append([row.get(c) for c in cols])
        sha = hashlib.sha256(_row_block_bytes(rows, delimiter=",")).hexdigest()
        per_sheet_records.append({
            "sheet_name": sheet_name,
            "figure_id": fig_id,
            "script": script or "",
            "source": per_sheet_sources.get(sheet_name, source or ""),
            "sha256": sha,
            "stamped_at": stamped,
            "description": desc,
            "n_rows": len(rows),
        })

    prov_ws = wb.create_sheet(title="Provenance")
    prov_cols = ["sheet_name", "figure_id", "script", "source", "sha256",
                 "stamped_at", "description", "n_rows"]
    prov_ws.append(prov_cols)
    # Make the Provenance header bold so the sheet reads as a key.
    for c in next(prov_ws.iter_rows(min_row=1, max_row=1)):
        c.font = Font(bold=True)
    for rec in per_sheet_records:
        prov_ws.append([rec[c] for c in prov_cols])

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)

    return {
        "path": str(path),
        "stamped_at": stamped,
        "sheets": per_sheet_records,
    }


def _format_banner(figure_id: str) -> str:
    """Turn a figure_id like 'figure_3_panel_b' or 'Figure 3' or
    'supplementary_figure_S1' into a human-readable banner string."""
    s = figure_id.strip()
    if not s:
        return ""
    # Already human-readable? Pass through.
    if s.startswith(("Figure", "Supplementary", "Backs")):
        return s if s.startswith("Backs") else f"Backs {s}"
    # snake_case → Title Case with the underscore-separated parts joined.
    parts = s.replace("-", "_").split("_")
    head = " ".join(p[:1].upper() + p[1:] for p in parts if p)
    return f"Backs {head}"


def _col_letter(n: int) -> str:
    """1 → 'A', 27 → 'AA', etc."""
    out = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        out = chr(ord("A") + rem) + out
    return out


# ---------- CLI --------------------------------------------------------------


def _read_rows(path: Union[str, os.PathLike], delimiter: str) -> List[Dict[str, Any]]:
    with open(path, encoding="utf-8", newline="") as fh:
        reader = csv.DictReader(fh, delimiter=delimiter)
        return list(reader)


def _cli() -> int:
    p = argparse.ArgumentParser(description="Write a figure-stamped supplementary table.")
    sub = p.add_subparsers(dest="cmd", required=True)

    common_prov = lambda x: (
        x.add_argument("--figure"),
        x.add_argument("--script"),
        x.add_argument("--source"),
        x.add_argument("--description"),
    )

    csv_p = sub.add_parser("csv", help="Write a CSV.")
    csv_p.add_argument("out")
    csv_p.add_argument("--rows-from", required=True, help="Source CSV to read rows from.")
    common_prov(csv_p)

    tsv_p = sub.add_parser("tsv", help="Write a TSV.")
    tsv_p.add_argument("out")
    tsv_p.add_argument("--rows-from", required=True, help="Source TSV / CSV to read rows from.")
    tsv_p.add_argument("--input-delimiter", default=",", help="Delimiter of --rows-from (default ',').")
    common_prov(tsv_p)

    xlsx_p = sub.add_parser("xlsx", help="Write a multi-sheet XLSX.")
    xlsx_p.add_argument("out")
    xlsx_p.add_argument("--sheet", action="append", required=True,
                        metavar="NAME=PATH",
                        help="Sheet name + source CSV (repeatable).")
    xlsx_p.add_argument("--input-delimiter", default=",")
    common_prov(xlsx_p)

    a = p.parse_args()
    if a.cmd == "csv":
        rows = _read_rows(a.rows_from, ",")
        info = write_csv(a.out, rows, figure=a.figure, script=a.script,
                         source=a.source, description=a.description)
    elif a.cmd == "tsv":
        rows = _read_rows(a.rows_from, a.input_delimiter)
        info = write_tsv(a.out, rows, figure=a.figure, script=a.script,
                         source=a.source, description=a.description)
    elif a.cmd == "xlsx":
        sheets: Dict[str, List[Dict[str, Any]]] = {}
        for spec in a.sheet:
            if "=" not in spec:
                raise SystemExit(f"--sheet expects NAME=PATH, got {spec!r}")
            name, path = spec.split("=", 1)
            sheets[name] = _read_rows(path, a.input_delimiter)
        info = write_xlsx(a.out, sheets, figure=a.figure, script=a.script,
                          source=a.source, description=a.description)
    else:
        raise SystemExit(2)
    print(info)
    return 0


if __name__ == "__main__":
    sys.exit(_cli())
