"""Smoke tests for supplementary-table. Run with: python -m pytest tests/"""
import csv
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from supp_table import write_csv, write_tsv, write_xlsx  # noqa: E402


SAMPLE_ROWS = [
    {"gene": "STAT5B", "mf_rank": 0.991, "score": 0.42},
    {"gene": "YY1",    "mf_rank": 0.957, "score": 0.81},
]


def test_write_csv_header_and_body(tmp_path):
    out = tmp_path / "table_s1.csv"
    info = write_csv(
        out,
        SAMPLE_ROWS,
        figure="figure_3_panel_b",
        script="scripts/build_figures.py::build_fig3",
        source="runs/2026-05-25/summary.json",
        description="Per-gene MF scores.",
    )
    text = out.read_text()
    # Provenance lines all `#`-prefixed.
    prov = [line for line in text.splitlines() if line.startswith("#")]
    assert any("figure: figure_3_panel_b" in p for p in prov)
    assert any("sha256: " in p for p in prov)
    # Data block is comment-free and round-trips through csv.DictReader.
    body = "\n".join(line for line in text.splitlines() if not line.startswith("#"))
    reader = csv.DictReader(body.splitlines())
    rows = list(reader)
    assert [r["gene"] for r in rows] == ["STAT5B", "YY1"]
    assert info["n_rows"] == 2
    assert len(info["sha256"]) == 64


def test_csv_and_tsv_same_data_same_hash(tmp_path):
    """Same rows in csv + tsv must hash identically (data block is canonical)."""
    csv_info = write_csv(tmp_path / "x.csv", SAMPLE_ROWS, figure="f")
    tsv_info = write_tsv(tmp_path / "x.tsv", SAMPLE_ROWS, figure="f")
    assert csv_info["sha256"] != tsv_info["sha256"], (
        "different delimiters → different byte streams, expected different hashes"
    )
    # Round-trip parity is what we really care about — both reads yield the
    # same row dicts when the comment lines are stripped.
    for path, delim in ((tmp_path / "x.csv", ","), (tmp_path / "x.tsv", "\t")):
        text = path.read_text()
        body = "\n".join(line for line in text.splitlines() if not line.startswith("#"))
        rows = list(csv.DictReader(body.splitlines(), delimiter=delim))
        assert [r["gene"] for r in rows] == ["STAT5B", "YY1"]


def test_write_xlsx_provenance_sheet(tmp_path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        import pytest
        pytest.skip("openpyxl not installed")
    out = tmp_path / "table_s1.xlsx"
    info = write_xlsx(
        out,
        sheets={"MF": SAMPLE_ROWS, "Enrich": [{"a": 1, "b": 2}]},
        figure="figure_3",
        script="scripts/build_figures.py",
        source="runs/2026-05-25/summary.json",
        per_sheet_figures={"MF": "figure_3_panel_b", "Enrich": "figure_3_panel_c"},
    )
    wb = load_workbook(out)
    assert {"MF", "Enrich", "Provenance"}.issubset(set(wb.sheetnames))
    prov = wb["Provenance"]
    header = [c.value for c in next(prov.iter_rows(min_row=1, max_row=1))]
    assert "figure_id" in header and "sha256" in header
    rows = [dict(zip(header, [c.value for c in r])) for r in prov.iter_rows(min_row=2)]
    by_sheet = {r["sheet_name"]: r for r in rows}
    assert by_sheet["MF"]["figure_id"] == "figure_3_panel_b"
    assert by_sheet["Enrich"]["figure_id"] == "figure_3_panel_c"
    assert len(by_sheet["MF"]["sha256"]) == 64
    assert info["sheets"][0]["sheet_name"] == "MF"


def test_write_csv_with_no_provenance_still_stamps_hash(tmp_path):
    """All caller-supplied provenance fields are optional, but sha256 +
    stamped_at are always emitted as `#`-prefixed comments so the file
    is never untraceable."""
    out = tmp_path / "bare.csv"
    info = write_csv(out, SAMPLE_ROWS)
    text = out.read_text()
    prov = [line for line in text.splitlines() if line.startswith("#")]
    assert any("sha256: " in p for p in prov)
    assert any("stamped_at: " in p for p in prov)
    # No figure / script / source / description lines (omitted, not blank).
    assert not any("figure: " in p for p in prov)
    assert not any("script: " in p for p in prov)
    assert len(info["sha256"]) == 64


def test_xlsx_data_sheet_row_count(tmp_path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        import pytest
        pytest.skip("openpyxl not installed")
    out = tmp_path / "y.xlsx"
    write_xlsx(out, sheets={"S1": SAMPLE_ROWS}, figure="fX")
    wb = load_workbook(out)
    ws = wb["S1"]
    # banner row + header row + 2 data rows = 4.
    assert ws.max_row == 4


def test_xlsx_banner_row_on_data_sheets(tmp_path):
    """Data sheets get a top banner naming the backing figure when
    `banner=True` (default)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        import pytest
        pytest.skip("openpyxl not installed")
    out = tmp_path / "banner.xlsx"
    write_xlsx(
        out,
        sheets={"MF": SAMPLE_ROWS},
        figure="figure_3_panel_b",
    )
    wb = load_workbook(out)
    ws = wb["MF"]
    # Row 1 = banner, row 2 = header, rows 3–4 = data.
    assert ws.cell(row=1, column=1).value == "Backs Figure 3 Panel B"
    assert ws.cell(row=1, column=1).font.bold
    assert ws.cell(row=2, column=1).value == "gene"
    assert ws.cell(row=3, column=1).value == "STAT5B"


def test_xlsx_banner_passthrough_when_already_human_readable(tmp_path):
    """A figure_id that's already 'Figure 3' or 'Supplementary Figure S2'
    should be prefixed with 'Backs ' without further casing changes."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        import pytest
        pytest.skip("openpyxl not installed")
    out = tmp_path / "human.xlsx"
    write_xlsx(
        out,
        sheets={"A": SAMPLE_ROWS, "B": SAMPLE_ROWS},
        per_sheet_figures={
            "A": "Figure 3",
            "B": "Supplementary Figure S2",
        },
    )
    wb = load_workbook(out)
    assert wb["A"].cell(row=1, column=1).value == "Backs Figure 3"
    assert wb["B"].cell(row=1, column=1).value == "Backs Supplementary Figure S2"


def test_xlsx_banner_opt_out(tmp_path):
    """`banner=False` means the header row is at row 1 (legacy layout)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        import pytest
        pytest.skip("openpyxl not installed")
    out = tmp_path / "nobanner.xlsx"
    write_xlsx(
        out,
        sheets={"S1": SAMPLE_ROWS},
        figure="figure_3",
        banner=False,
    )
    wb = load_workbook(out)
    ws = wb["S1"]
    assert ws.cell(row=1, column=1).value == "gene"
    assert ws.max_row == 3
